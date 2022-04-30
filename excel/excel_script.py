import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook

# import xlsxwriter

# df = pd.DataFrame({'States':['California', 'Florida', 'Montana', 'Colorodo', 'Washington', 'Virginia'],
#                     'Capitals':['Sacramento', 'Tallahassee', 'Helena', 'Denver', 'Olympia', 'Richmond'],
#                     'Population':['508529', '193551', '32315', '619968', '52555', '227032']})
# print(df)

# df.to_excel('./states.xlsx')
# df.to_excel('./states.xlsx', sheet_name="first_sheet", index=False)


################################################################################################
###################                    Create many sheets               ########################
################################################################################################

# income1 = pd.DataFrame({'Names': ['Stephen', 'Camilla', 'Tom'], 
#                         'Salary': [100000, 70000, 60000]})
# income2 = pd.DataFrame({'Names': ['Pete', 'April', 'Marty'],
#                         'Salary':[120000, 110000, 50000]})
# income3 = pd.DataFrame({'Names': ['Victor', 'Victoria', 'Jennifer'],
#                         'Salary':[75000, 90000, 40000]})
# income_sheets = {'income1': income1, 'income2': income2, 'income3': income3}
# writer = pd.ExcelWriter('./income.xlsx', engine='xlsxwriter')

# for sheet_name in income_sheets.keys():
#     income_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

# writer.save()


###############################################################################################
##################           Read an excel file with  many sheets            ##################
###############################################################################################

# students_grades = pd.DataFrame({'Student Name':['Jennifer', 'Stephen', 'Alfred', 'Anastasia', 'Dicken', 'Davy'],
#                     'Grade':[6, 8, 2, 10, 10, 8],
#                     'Marks Obtained':[81.1, 140.0, 27.5, 380.0, 450.0, 300.0], 
#                     'Total Markes':[100, 200, 50, 500, 500, 450], 
#                     'Percentage':['81%', '70%', '55%', '76%', '90%', '75%']})
# path  = os.getcwd()
# if os.path.isfile(os.path.join(path, 'grades.xlsx')):
#     os.remove('grades.xlsx')
# students_grades.to_excel('./grades.xlsx', sheet_name='Grades', index=False)        

# Reading the excel file
# df = pd.read_excel('./grades.xlsx', sheet_name='Grades', index_col='Grade')
# print(df)



################################################################################################
##################         Reading specific column in an excel file            #################
################################################################################################


# cols = [0, 2, 4] # let us specify columns to display in the dataframe
# df = pd.read_excel('./grades.xlsx', sheet_name='Grades', usecols=cols)
# print(df.head())


################################################################################################
##################         write a new sheet to an existing excel file         #################
################################################################################################


## OVerwriting the old sheet
# States = pd.DataFrame({'States':['California', 'Florida', 'Montana', 'Colorodo', 'Washington', 'Virginia'],
#                     'Capitals':['Sacramento', 'Tallahassee', 'Helena', 'Denver', 'Olympia', 'Richmond'],
#                     'Population':['508529', '193551', '32315', '619968', '52555', '227032']})

# writer = pd.ExcelWriter('grades.xlsx', engine='xlsxwriter')
# States.to_excel(writer, sheet_name='states', index=False)
# writer.save()
# writer.close()


# ## add a new sheet to an existing excel file without losing the original 
# States = pd.DataFrame({'States':['California', 'Florida', 'Montana', 'Colorodo', 'Washington', 'Virginia'],
#                     'Capitals':['Sacramento', 'Tallahassee', 'Helena', 'Denver', 'Olympia', 'Richmond'],
#                     'Population':['508529', '193551', '32315', '619968', '52555', '227032']})

# grades = load_workbook('./grades.xlsx')
# writer = pd.ExcelWriter('./grades.xlsx', engine='openpyxl')
# writer.book = grades
# States.to_excel(writer, sheet_name='states', index=False)
# writer.save()
# writer.close()

### Change sheet title
my_file=openpyxl.load_workbook("./grades.xlsx")
my_file_sheet= my_file['Grades']
my_file_sheet.title = 'Student_grades'
my_file.save("grades.xlsx")