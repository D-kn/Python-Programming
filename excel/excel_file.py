# from codecs import ignore_errors
# from turtle import shearfactor
# from unicodedata import category
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
import json

import warnings 
warnings.filterwarnings('ignore')

################################################################################################
##################               Read an Excel Spreadsheet                     #################
################################################################################################

os.chdir(r'C:\Users\p114215\Documents\Internship\coding\excel')
print(os.getcwd())

# workbook = load_workbook(filename="reviews-sample.xlsx") # open the spreadsheet
# print(workbook.sheetnames)

# sheet = workbook.active
# print(sheet)
# print(sheet.title)
# print(sheet["A1"])
# print(sheet["A1"].value)
# print(sheet["F10"].value) # get the value in the current cell
# print(sheet.cell(row=10, column=6).value)
# print(sheet["A:B"]) 

# # generators
# for column in sheet.iter_cols(min_row=1, 
#                               max_row=2, 
#                               min_col=1, 
#                               max_col=3):
#     print(column)
# print()
# for row in sheet.iter_rows(min_row=1, 
#                            max_row=2, 
#                            min_col=1, 
#                            max_col=3):
#     print(row)

# for value in sheet.iter_rows(min_row=1, 
#                             max_row=2, 
#                             min_col=1, 
#                             max_col=4,
#                             values_only=True):
#     print(value)


#----------------------------------------------------------------------------
##-------------------#          JSON Files             #-------------------##

# workbook = load_workbook(filename="reviews-sample.xlsx") 
# sheet = workbook.active

# products = {}

# # return cell values
# for value in sheet.iter_rows(max_row=2, 
#                               min_col=4, 
#                               max_col=7, 
#                               values_only=True):
#     product_id = value[0]                              
#     product = {
#         "parent":value[1],
#         "title":value[2],
#         "category":value[3]
#     }                             
#     products[product_id] = product

# print(json.dumps(products))


#----------------------------------------------------------------------------
##-------------------#        Appending new data       #-------------------##

# workbook = load_workbook(filename="grades.xlsx")
# sheet = workbook.active

# print(sheet.title)

# # write what I want
# sheet["A8"]="Noel"
# sheet["B8"]=7
# sheet["C8"]=266
# sheet["D8"]=350
# sheet["E8"]="90%"

# # Save the spreadsheet
# workbook.save(filename="grades.xlsx")

# def print_rows():
#     for row in sheet.iter_rows(values_only=True):
#         print(row)

# print(print_rows())

#----------------------------------------------------------------------------
##-------------------#    Managing rows and columns    #-------------------##
















################################################################################################
##################              Convert Data into python Classes               #################
################################################################################################
